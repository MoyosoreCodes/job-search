from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from config.config import AppConfig

class ExcelHandler:
    def __init__(self, results_root: Path):
        self.results_root = Path(results_root)

    def _dated_results_dir(self) -> Path:
        d = datetime.now().strftime("%b_%d_%Y")
        return self.results_root / d

    def write(self, jobs: list) -> Path:
        out_dir = self._dated_results_dir()
        out_dir.mkdir(parents=True, exist_ok=True)
        name = datetime.now().strftime("job_search_result_%Y%m%d_%H%M%S.xlsx")
        fpath = out_dir / name
        df = pd.DataFrame(jobs)
        if "cover_letter_generated" not in df.columns:
            df["cover_letter_generated"] = ""
        if "application_status" not in df.columns:
            df["application_status"] = ""
        df.to_excel(fpath, sheet_name=AppConfig.SHEET_NAME, index=False)
        self._apply_validation(fpath)
        return fpath

    def _apply_validation(self, fpath: Path) -> None:
        try:
            wb = load_workbook(fpath)
            ws = wb[AppConfig.SHEET_NAME]
            headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
            col_idx = headers.get("application_status") or headers.get("application status")
            if not col_idx:
                wb.save(fpath)
                return
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            opts = ",".join(AppConfig.APPLICATION_STATUS_OPTIONS)
            dv = DataValidation(type="list", formula1=f'"{opts}"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.ranges.append(f"{col_letter}2:{col_letter}{ws.max_row}")
            wb.save(fpath)
        except Exception:
            pass
